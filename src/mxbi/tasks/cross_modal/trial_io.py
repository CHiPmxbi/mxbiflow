import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from mxbi.utils.logger import logger

from mxbi.tasks.cross_modal.trial_schema import Trial


@dataclass
class TrialStore:
    trials: list[Trial]
    by_subject: dict[str, list[Trial]]
    indices: dict[str, int]

    @classmethod
    def from_trials(cls, trials: list[Trial]) -> "TrialStore":
        trials_sorted = sorted(trials, key=lambda t: (t.subject_id, t.trial_number))
        by_subject: dict[str, list[Trial]] = {}
        for t in trials_sorted:
            by_subject.setdefault(t.subject_id, []).append(t)
        return cls(trials_sorted, by_subject, {})

    def trials_for_subject(self, subject_id: str) -> list[Trial]:
        return self.by_subject.get(subject_id, [])

    def next_for_subject(self, subject_id: str) -> tuple[Trial, int]:
        subject_trials = self.trials_for_subject(subject_id)
        if not subject_trials:
            raise ValueError(f"No trials for subject_id '{subject_id}'")

        idx = self.indices.get(subject_id, 0)
        i = idx % len(subject_trials)
        return subject_trials[i], i

    def advance(self, subject_id: str, last_index: int) -> None:
        self.indices[subject_id] = last_index + 1


_TRIAL_STORE_CACHE: dict[tuple[Path, Path | None], TrialStore] = {}


def read_trials(path: Path, base_dir: Path | None = None) -> TrialStore:
    if not path.exists():
        raise FileNotFoundError(f"Trial file not found: {path}")

    key = (path.resolve(), base_dir.resolve() if base_dir else None)
    cached = _TRIAL_STORE_CACHE.get(key)
    if cached is not None:
        return cached

    ext = path.suffix.lower()
    if ext == ".csv":
        records = _read_csv(path)
    elif ext == ".json":
        records = _read_json(path)
    elif ext in {".xlsx"}:
        records = _read_xlsx(path)
    else:
        raise ValueError(f"Unsupported trial file format: {path.suffix}")

    trials: list[Trial] = []
    for rec in records:
        t = Trial.model_validate(rec)
        _assert_paths_exist(t, base_dir)
        trials.append(t)

    logger.info("Loaded %d trials from %s", len(trials), path)
    store = TrialStore.from_trials(trials)
    _TRIAL_STORE_CACHE[key] = store
    return store


def _read_csv(path: Path) -> Iterable[dict]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def _read_json(path: Path) -> Iterable[dict]:
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        for rec in data:
            yield rec
    elif isinstance(data, dict):
        if "trials" in data and isinstance(data["trials"], list):
            for rec in data["trials"]:
                yield rec
        else:
            raise ValueError("JSON must be a list of trials or { 'trials': [...] }")
    else:
        raise ValueError("Malformed JSON trial file")


def _read_xlsx(path: Path) -> Iterable[dict]:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("Reading .xlsx requires openpyxl. Install it or provide CSV/JSON.")
    wb = load_workbook(filename=str(path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) for h in rows[0]]
    for r in rows[1:]:
        rec: dict[str, object] = {}
        for k, v in zip(header, r):
            rec[k] = v
        yield rec


def _assert_paths_exist(t: Trial, base_dir: Path | None) -> None:
    ap = t.audio_path_obj(base_dir)
    lp = t.left_image_path_obj(base_dir)
    rp = t.right_image_path_obj(base_dir)
    missing = [str(p) for p in (ap, lp, rp) if not p.exists()]
    if missing:
        raise FileNotFoundError(
            "Stimulus paths do not exist for "
            f"trial_id={t.trial_id}, subject_id={t.subject_id}: {missing}"
        )
